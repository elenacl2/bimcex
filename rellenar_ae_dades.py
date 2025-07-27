import os
import openpyxl
from openpyxl.styles import Alignment
from gbxml_reader import leer_gbxml

def generar_excel_datos_ae(ruta_xml):
    espacios, superficies, _ = leer_gbxml(ruta_xml)
    wb = openpyxl.Workbook()
    hoja_por_defecto = wb.active
    hoja_por_defecto.title = "Resumen"
    hoja_util_creada = False

    hojas = [
        "3.1 Coberta",
        "3.2 Mur",
        "3.3 Sòl",
        "3.4 Partició interior",
        "3.5 Buit ò lluernari",
        "3.6 Ponts tèrmics"
    ]

    encabezados_generales = [
        "Nom", "Zona", "Superfície (m2)", "Orientació (graus)",
        "Transmitància tèrmica (W/m2K)", "Composició"
    ]

    mapa_hoja_tipos = {
        "3.1 Coberta": ["Roof"],
        "3.2 Mur": ["ExteriorWall"],
        "3.3 Sòl": ["SlabOnGrade", "UndergroundSlab"],
        "3.4 Partició interior": ["InteriorWall"]
    }

    for hoja in hojas[:4]:
        ws = wb.create_sheet(title=hoja)
        ws.append(encabezados_generales)
        tipos = mapa_hoja_tipos[hoja]
        datos_encontrados = False

        for superficie in superficies:
            if superficie["tipo"] in tipos:
                nombre = superficie.get("construccion") or "Sense nom"
                zona = ", ".join(superficie.get("espacios_adyacentes", [])) or "Sense espai"
                area = round(superficie.get("area", 0), 2)
                azimuth = superficie.get("azimuth")
                orientacion = round(float(azimuth), 2) if azimuth is not None else ""
                transmitancia = superficie.get("transmitancia", "")
                composicion = superficie.get("construccion", "")

                fila = [nombre, zona, area, orientacion, transmitancia, composicion]
                ws.append(fila)
                datos_encontrados = True

        if datos_encontrados:
            hoja_util_creada = True
        else:
            wb.remove(ws)

    # Hoja 3.5 Buit ò lluernari
    ws_buits = wb.create_sheet(title="3.5 Buit ò lluernari")
    encabezados_buits = [
        "Tipus de buit", "Nom", "Zona", "Superfície (m2)",
        "Orientació (graus)", "Transmitància vidre (W/m2K)", "Composició"
    ]
    ws_buits.append(encabezados_buits)
    found_openings = False

    for superficie in superficies:
        zona = ", ".join(superficie.get("espacios_adyacentes", [])) or "Sense espai"
        azimuth = superficie.get("azimuth")
        orientacion = round(float(azimuth), 2) if azimuth is not None else ""

        for idx, opening in enumerate(superficie.get("openings", [])):
            tipo = opening.get("tipo", "").lower()
            if tipo not in ["window", "skylight"]:
                continue

            found_openings = True
            tipus_buit = "Finestra" if tipo == "window" else "Llucernari"
            nom = opening.get("id") or f"{tipus_buit}_{idx + 1}"
            area = round(opening.get("area", 0), 2)
            transmitancia = opening.get("transmitancia_vidre", "")
            composicio = opening.get("tipus_vidre", "")

            fila = [tipus_buit, nom, zona, area, orientacion, transmitancia, composicio]
            ws_buits.append(fila)

    if found_openings:
        hoja_util_creada = True
    else:
        wb.remove(ws_buits)

    # Hoja 3.6 Ponts tèrmics
    ws_ponts = wb.create_sheet(title="3.6 Ponts tèrmics")
    encabezados_ponts = [
        "Tipus de pont tèrmic", "Nom", "Zona", "Longitud (m)",
        "Psi (W/mK)", "Descripció", "Observacions"
    ]
    ws_ponts.append(encabezados_ponts)
    hoja_util_creada = True  # Se considera válida aunque esté vacía

    if hoja_util_creada:
        wb.remove(hoja_por_defecto)
    else:
        hoja_por_defecto.append(["⚠️ No se encontraron datos válidos en el archivo gbXML."])

    # Formato
    for ws in wb.worksheets:
        for col in ws.columns:
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical="center")

    nombre_base = os.path.splitext(os.path.basename(ruta_xml))[0]
    ruta_salida = os.path.join(os.path.dirname(ruta_xml), f"{nombre_base}_datos_ae.xlsx")
    wb.save(ruta_salida)
    print(f"✅ Archivo generado correctamente: {ruta_salida}")

